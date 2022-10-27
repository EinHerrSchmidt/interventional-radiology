from openpyxl import load_workbook
from enum import Enum
import csv
import datetime
from os import listdir
import logging

from exceptions import MainSheetNotFound
from model import Procedure


class ExcelIndex(Enum):
    DATE_TIME = 0
    GENDER = 1
    SERVICE = 2
    CLINICAL_QUESTION = 3
    DIAGNOSTIC = 4
    ORDER_STATUS = 5
    MEDICAL_DOCTOR = 6
    HOSPITAL_WARD = 7
    REQUEST_NUMBER = 8
    ACCESS_NUMBER = 9
    PATIENT_ID = 10
    EPISODE_NUMBER = 11


def add_header(file_name):
    header = ["DATE", "TIME", "GENDER", "BIRTH_DATE", "SERVICE", "CLINICAL_QUESTION", "DIAGNOSTIC", "ORDER__STATUS",
              "MEDICAL_DOCTOR", "HOSPITAL_WARD", "REQUEST_NUMBER", "ACCESS_NUMBER", "PATIENT_ID", "EPISODE_NUMBER"]
    with open(file_name, 'w', newline='') as csvfile:
        file_writer = csv.writer(
            csvfile, delimiter='\t', quotechar='"', quoting=csv.QUOTE_MINIMAL)
        file_writer.writerow(header)


def write_rows(procedures, file_name):
    with open(file_name, mode='a', newline='') as csvfile:
        file_writer = csv.writer(csvfile, delimiter='\t', quotechar='"', quoting=csv.QUOTE_MINIMAL)
        for procedure in procedures:
            file_writer.writerow(procedure.to_list())

def find_main_sheet(workbook):
    sheet_names = workbook.sheetnames
    for sheet_name in sheet_names:
        workbook.active = workbook[sheet_name]
        # we check for the right header
        if list(workbook.active.values)[0][0] == "Data inizio":
            return sheet_name
    raise MainSheetNotFound("Main sheet not found")


def load_procedures(xlsx_file_name):
    wb = load_workbook(xlsx_file_name)
    sheet_names = wb.sheetnames
    main_sheet = find_main_sheet(wb)
    active_sheet = wb[main_sheet]
    header = True
    procedures = []
    for row in list(active_sheet.values):
        if header:
            header = False
            continue
        date_and_time = row[ExcelIndex.DATE_TIME.value].strip().split("\n")
        date_tokens = date_and_time[0].split("/")
        time_tokens = date_and_time[1].split(":")
        procedure_date_time = datetime.datetime(year=int(date_tokens[2]),
                                                month=int(date_tokens[1]),
                                                day=int(date_tokens[0]),
                                                hour=int(time_tokens[0]),
                                                minute=int(time_tokens[1]))

        gender_and_birth = row[ExcelIndex.GENDER.value].strip().split(" ")
        gender = gender_and_birth[0][0]
        birth_tokens = gender_and_birth[1].split("/")
        birth_date = datetime.datetime(year=int(birth_tokens[2]),
                                       month=int(birth_tokens[1]),
                                       day=int(birth_tokens[0]))

        service = row[ExcelIndex.SERVICE.value].strip().replace("\n", "")
        clinical_question = row[ExcelIndex.CLINICAL_QUESTION.value].strip().replace("\n", "")
        diagnostic = row[ExcelIndex.DIAGNOSTIC.value].strip().replace("\n", "")
        order_status = row[ExcelIndex.ORDER_STATUS.value].strip().replace("\n", "")
        medical_doctor = row[ExcelIndex.MEDICAL_DOCTOR.value].strip().replace("\n", "")
        hospital_ward = row[ExcelIndex.HOSPITAL_WARD.value].strip().replace("\n", "")
        request_number = row[ExcelIndex.REQUEST_NUMBER.value].strip().replace("\n", "")
        access_number = row[ExcelIndex.ACCESS_NUMBER.value].strip().replace("\n", "")
        patient_id = row[ExcelIndex.PATIENT_ID.value].strip().replace("\n", "")
        try:
            episode_number = str(
                int(row[ExcelIndex.EPISODE_NUMBER.value])).strip().replace("\n", "")
        except ValueError:
            episode_number = row[ExcelIndex.EPISODE_NUMBER.value].strip().replace("\n", "")

        procedure = Procedure(procedure_date_time.date(),
                              procedure_date_time.time(),
                              gender, birth_date.date(),
                              service,
                              clinical_question,
                              diagnostic,
                              order_status,
                              medical_doctor,
                              hospital_ward,
                              request_number,
                              access_number,
                              patient_id,
                              episode_number,
                              None,
                              None)

        procedures.append(procedure)
    return procedures

current_date_time = datetime.datetime.now()
current_date_tokens = str(current_date_time.date()).split("-")
current_time_tokens = str(current_date_time.time()).split(":")

logging.basicConfig(filename="./logs/" + "".join(current_date_tokens) + "_" + "".join(list(map(str, list(map(int, list(map(float, current_time_tokens))))))[0:3]) + ".log", encoding="utf-8", level=logging.DEBUG)
add_header("procedures.csv")
procedures = []
for file_name in listdir("./xlsx"):
    try:
        procedures.extend(load_procedures("./xlsx/" + file_name))
    except MainSheetNotFound:
        logging.warning("File " + file_name + " was skipped since it has no main sheet.")
        continue
write_rows(procedures, "procedures.csv")
# logging.info("File " + file_name + " loaded.")