import csv
from enum import Enum
from openpyxl import load_workbook
from exceptions import MainSheetNotFound
import datetime
from os import listdir
import logging
import re


class Procedure():
    def __init__(self, date, time, patient_gender, patient_birth_date, service, clinical_question, diagnostic, order_status, medical_doctor, hospital_ward, request_number, access_number, patient_id, episode_number, or_entrance, or_leave, or_occupation):
        self.date = date
        self.time = time
        self.patient_gender = patient_gender
        self.patient_birth_date = patient_birth_date
        self.service = service
        self.clinical_question = clinical_question
        self.diagnostic = diagnostic
        self.order_status = order_status
        self.medical_doctor = medical_doctor
        self.hospital_ward = hospital_ward
        self.request_number = request_number
        self.access_number = access_number
        self.patient_id = patient_id
        self.episode_number = episode_number
        self.or_entrance = or_entrance
        self.or_leave = or_leave
        self.or_occupation = or_occupation

    def to_list(self):
        return [self.date,
                self.time,
                self.patient_gender,
                self.patient_birth_date,
                self.service,
                self.clinical_question,
                self.diagnostic,
                self.order_status,
                self.medical_doctor,
                self.hospital_ward,
                self.request_number,
                self.access_number,
                self.patient_id,
                self.episode_number,
                self.or_entrance,
                self.or_leave,
                self.or_occupation]

    def __str__(self):
        return str(self.date) + " " + str(self.time) + " " + self.patient_gender + " " + str(self.patient_birth_date) + " " + self.service + " " + self.clinical_question + " " + self.diagnostic + " " + self.order_status + " " + self.medical_doctor + " " + self.hospital_ward + " " + self.request_number + " " + self.access_number + " " + self.patient_id + " " + self.episode_number + " " + str(self.or_entrance) + " " + str(self.or_leave)


class CSVWriter:
    def __init__(self, file_name):
        self.file_name = file_name

    def add_header(self, header, file_name):
        with open(file_name, 'w', newline='') as csvfile:
            file_writer = csv.writer(
                csvfile, delimiter='\t', quotechar='"', quoting=csv.QUOTE_MINIMAL)
            file_writer.writerow(header)

    def write_rows(self, lists, file_name):
        with open(file_name, mode='a', newline='') as csvfile:
            file_writer = csv.writer(csvfile, delimiter='\t', quotechar='"', quoting=csv.QUOTE_MINIMAL)
            file_writer.writerows(lists)

    def write(self, header, lists):
        self.add_header(header, self.file_name)
        self.write_rows(lists, self.file_name)


class ExcelIndex(Enum):
    DATE_TIME = 0
    GENDER = 1
    SERVICE = 2
    CLINICAL_QUESTION = 3
    DIAGNOSTIC = 4
    ORDER_STATUS = 5
    MEDICAL_DOCTOR = 6
    HOSPITAL_WARD = 7
    REQUEST_NUMBER = 8
    ACCESS_NUMBER = 9
    PATIENT_ID = 10
    EPISODE_NUMBER = 11


class ExcelLoader:
    def __init__(self, xlsx_dir):
        self.xlsx_dir = xlsx_dir

    def find_main_sheet(self, workbook):
        sheet_names = workbook.sheetnames
        for sheet_name in sheet_names:
            workbook.active = workbook[sheet_name]
            # we check for the right header
            if list(workbook.active.values)[0][0] == "Data inizio":
                return sheet_name
        raise MainSheetNotFound("Main sheet not found")

    def load_procedures(self, xlsx_file_name):
        wb = load_workbook(xlsx_file_name)
        main_sheet = self.find_main_sheet(wb)
        active_sheet = wb[main_sheet]
        header = True
        procedures = []
        for row in list(active_sheet.values):
            if header:
                header = False
                continue
            date_and_time = row[ExcelIndex.DATE_TIME.value].strip().split("\n")
            date_tokens = date_and_time[0].split("/")
            time_tokens = date_and_time[1].split(":")
            procedure_date_time = datetime.datetime(year=int(date_tokens[2]),
                                                    month=int(date_tokens[1]),
                                                    day=int(date_tokens[0]),
                                                    hour=int(time_tokens[0]),
                                                    minute=int(time_tokens[1]))

            gender_and_birth = row[ExcelIndex.GENDER.value].strip().split(" ")
            gender = gender_and_birth[0][0]
            birth_tokens = gender_and_birth[1].split("/")
            birth_date = datetime.datetime(year=int(birth_tokens[2]),
                                           month=int(birth_tokens[1]),
                                           day=int(birth_tokens[0]))

            service = row[ExcelIndex.SERVICE.value].strip().replace("\n", "")
            clinical_question = row[ExcelIndex.CLINICAL_QUESTION.value].strip().replace("\n", "")
            diagnostic = row[ExcelIndex.DIAGNOSTIC.value].strip().replace("\n", "")
            order_status = row[ExcelIndex.ORDER_STATUS.value].strip().replace("\n", "")
            medical_doctor = row[ExcelIndex.MEDICAL_DOCTOR.value].strip().replace("\n", "")
            hospital_ward = row[ExcelIndex.HOSPITAL_WARD.value].strip().replace("\n", "")
            request_number = row[ExcelIndex.REQUEST_NUMBER.value].strip().replace("\n", "")
            access_number = row[ExcelIndex.ACCESS_NUMBER.value].strip().replace("\n", "")
            patient_id = row[ExcelIndex.PATIENT_ID.value].strip().replace("\n", "")
            try:
                episode_number = str(
                    int(row[ExcelIndex.EPISODE_NUMBER.value])).strip().replace("\n", "")
            except ValueError:
                episode_number = row[ExcelIndex.EPISODE_NUMBER.value].strip().replace("\n", "")

            procedure = Procedure(procedure_date_time.date(),
                                  procedure_date_time.time(),
                                  gender, birth_date.date(),
                                  service,
                                  clinical_question,
                                  diagnostic,
                                  order_status,
                                  medical_doctor,
                                  hospital_ward,
                                  request_number,
                                  access_number,
                                  patient_id,
                                  episode_number,
                                  None,
                                  None,
                                  None)

            procedures.append(procedure)
        
        sheets = wb.sheetnames
        for s in sheets:
            if(s != main_sheet):
                active_sheet = wb[s] # block sheet
                or_entrance = ""
                or_leave = ""
                or_occupation = ""
                for row in list(active_sheet.values):
                    entrance_match = re.search("^INGRESSO IN SALA.*- (.*)", row[0].upper())
                    leave_match = re.search("^USCITA DALLA SALA.*- (.*) PERMANENZA.*: (.*)", row[0].upper())
                    if(entrance_match):
                        or_entrance = entrance_match.group(1)
                    if(leave_match):
                        or_leave = leave_match.group(1)
                        or_occupation = leave_match.group(2)
                for p in procedures:
                    if(p.episode_number == s):
                        p.or_entrance = or_entrance
                        p.or_leave = or_leave
                        p.or_occupation = or_occupation

        return procedures

    def get_procedures(self):
        current_date_time = datetime.datetime.now()
        current_date_tokens = str(current_date_time.date()).split("-")
        current_time_tokens = str(current_date_time.time()).split(":")

        logging.basicConfig(filename="./logs/"
                            + "".join(current_date_tokens)
                            + "_"
                            + "".join(list(map(str, list(map(int, list(map(float, current_time_tokens))))))[0:3])
                            + ".log", encoding="utf-8", level=logging.DEBUG)
        procedures = []
        for file_name in listdir(self.xlsx_dir):
            try:
                procedures.extend(self.load_procedures(self.xlsx_dir
                                                       + "/"
                                                       + file_name))
            except MainSheetNotFound:
                logging.warning("File "
                                + file_name
                                + " was skipped since it has no main sheet.")
                continue
        return procedures
